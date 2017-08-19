# coding:utf-8
import os
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime
import sys

# 文件名 - 时间戳的准备
# tempFileData = {}
currentDay = datetime.datetime.now()
yesterday = currentDay + datetime.timedelta(days=-1)
# tempFileData[1] = yesterday.strftime("%Y/%m/%d")
# tempFileData[1] = currentDay.strftime("%Y/%m/%d")

# ------config-star-------
# 打开临时文件
# tempFilePath = "/data/select/alic_sjdshuju_" + yesterday.strftime("%Y-%m-%d") + ".xls"
tempFilePath = "/data/select/sjdshuju_" + yesterday.strftime("%Y-%m-%d") + ".xls"
sourceFilePath = "/data/select/手机端数据-" + yesterday.strftime("%Y%m%d") + ".xlsx"
targetFilePath = "/data/select/手机端数据-" + currentDay.strftime("%Y%m%d") + ".xlsx"
# ------config-end-------

# 判断文件是否存在
if not os.path.exists(tempFilePath):
    print tempFilePath + "文件不存在"
    print "Sorry! the task had stop~~~"
    sys.exit(0)

if not os.path.exists(sourceFilePath):
    print sourceFilePath + "文件不存在"
    print "Sorry! the task had stop~~~"
    sys.exit(0)

# 读取临时文件的数据
tempFile = open(tempFilePath, 'r')
# print tempFileData[1]
# index = 2
# while True:
#     line = tempFile.readline()
#     if line:
#         tempFileData[index] = line.strip()
#         # print tempFileData[index]
#         index += 1
#     else:
#         break
# tempFile.close()
tempFileData = tempFile.readline().strip().split('\t')
tempFileData.insert(0, yesterday.strftime("%Y/%m/%d"))
print tempFileData

# 加载目标文件并追加
wb = load_workbook(sourceFilePath)

# 打开默认的工作表
default_ws = wb.active

# 获取工作的行数
taskRow = len(list(default_ws.rows)) + 1
# print taskRow

# print len(default_ws.columns)

# 写入数据
print tempFileData
cell01 = default_ws.cell(row=taskRow, column=1).value = tempFileData[0]
cell02 = default_ws.cell(row=taskRow, column=2).value = tempFileData[1]
cell03 = default_ws.cell(row=taskRow, column=3).value = round(float(tempFileData[2]), 1)
cell04 = default_ws.cell(row=taskRow, column=4).value = tempFileData[3]
cell05 = default_ws.cell(row=taskRow, column=5).value = round(float(tempFileData[4]), 8)
cell06 = default_ws.cell(row=taskRow, column=6).value = round(float(tempFileData[5]), 8)
# cell01.value = tempFileData[1]
# cell02.value = tempFileData[2]
# cell03.value = tempFileData[3]
# cell04.value = tempFileData[4]
# cell05.value = tempFileData[5]
# cell06.value = tempFileData[6]

# 表格样式
# 遍历单元格文本居中
for row in default_ws.iter_rows('A1:F%d' % taskRow):
    for cell in row:
        # print cell
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 保存数据
print tempFileData
wb.save(targetFilePath)

print "the task is successful~~~"
