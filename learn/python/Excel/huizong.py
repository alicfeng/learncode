# coding:utf-8
import os
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime
import sys

# 文件名 - 时间戳的准备
# tempFileData = {}
currentDay = datetime.datetime.now()
yesterday = currentDay + datetime.timedelta(days=-1)
dateData01 = currentDay.strftime("%Y/%m/%d")
dateData02 = yesterday.strftime("%Y-%m-%d")

# 上一天的文件与新文路径
sourceFilePath = "/data/select/运营" + yesterday.strftime("%Y%m%d") + ".xlsx"
targetFilePath = "/data/select/运营" + currentDay.strftime("%Y%m%d") + ".xlsx"

# 渣文件
deng_lu = "/data/baobiao/" + dateData02 + "/denglurenshu.txt"
zhu_ce = "/data/baobiao/" + dateData02 + "/zhuceyonghu.txt"
money_er = "/data/baobiao/" + dateData02 + "/fufeiyonghu.txt"
pay_for = "/data/baobiao/" + dateData02 + "/fufeijinger.txt"
select_file = "/data/baobiao/" + dateData02 + "/alic_select.temp"

# 判断文件是否存在
if not os.path.exists(deng_lu):
    print deng_lu + "文件不存在"
    print "Sorry! the task had stop~~~"
    sys.exit(0)

if not os.path.exists(money_er):
    print money_er + "文件不存在"
    print "Sorry! the task had stop~~~"
    sys.exit(0)

if not os.path.exists(pay_for):
    print pay_for + "文件不存在"
    print "Sorry! the task had stop~~~"
    sys.exit(0)

if not os.path.exists(zhu_ce):
    print zhu_ce + "文件不存在"
    print "Sorry! the task had stop~~~"
    sys.exit(0)

if not os.path.exists(sourceFilePath):
    print sourceFilePath + "文件不存在"
    print "Sorry! the task had stop~~~"
    sys.exit(0)

if not os.path.exists(select_file):
    print select_file + "文件不存在"
    print "Sorry! the task had stop~~~"
    sys.exit(0)

# get data
print "源数据->"
# 登陆数据
deng_lu_temp_file = open(deng_lu, 'r')
deng_lu_temp_data = deng_lu_temp_file.readline()
deng_lu_data = deng_lu_temp_data.strip().split('\t')
print deng_lu_data

# 充值用户数据
money_er_temp_file = open(money_er, 'r')
money_er_temp_data = money_er_temp_file.readline()
money_er_data = money_er_temp_data.strip().split('\t')
print money_er_data

# 充值金额数据
pay_for_temp_file = open(pay_for, 'r')
pay_for_temp_data = pay_for_temp_file.readline()
pay_for_data = pay_for_temp_data.strip().split('\t')
print pay_for_data

# 注册数据
zhu_ce_temp_file = open(zhu_ce, 'r')
zhu_ce_temp_data = zhu_ce_temp_file.readline()
zhu_ce_data = zhu_ce_temp_data.strip().split('\t')
print zhu_ce_data

wb = load_workbook(filename=sourceFilePath)
# default_ws = wb.active

# ----------deng_lu start-----------
deng_lu_ws = wb.worksheets[1]
# 操作行数的位置
deng_lu_taskRow = len(deng_lu_ws.rows) + 1
deng_lu_taskColumn = len(deng_lu_ws.columns)

# print deng_lu_taskRow
# print deng_lu_taskColumn
# 转int
for index in range(0, len(deng_lu_data)):
    deng_lu_data[index] = float(deng_lu_data[index])

# 总和
for index in range(0, len(deng_lu_data)):
    deng_lu_data[len(deng_lu_data) - 2] += int(deng_lu_data[index])
# last value
deng_lu_data[len(deng_lu_data) - 1] = deng_lu_data[1] + deng_lu_data[2]

# 写数据
deng_lu_ws.cell(row=deng_lu_taskRow, column=1).value = dateData01
for index in range(0, len(deng_lu_data)):
    deng_lu_ws.cell(row=deng_lu_taskRow, column=index + 3).value = deng_lu_data[index]
    deng_lu_ws.cell(row=deng_lu_taskRow, column=index + 3).alignment = Alignment(horizontal='center', vertical='center')

# ----------deng_lu end-----------

# ----------zhu_ce start-----------
zhu_ce_ws = wb.worksheets[2]
# 操作行数的位置
zhu_ce_taskRow = len(zhu_ce_ws.rows) + 1
zhu_ce_taskColumn = len(zhu_ce_ws.columns)

# print zhu_ce_taskRow
# print zhu_ce_taskColumn
# 转int
for index in range(0, len(zhu_ce_data)):
    zhu_ce_data[index] = int(zhu_ce_data[index])

# 总和
for index in range(0, len(zhu_ce_data)):
    zhu_ce_data[len(zhu_ce_data) - 2] += int(zhu_ce_data[index])
# last value
zhu_ce_data[len(zhu_ce_data) - 1] = zhu_ce_data[1] + zhu_ce_data[2]

# 写数据
zhu_ce_ws.cell(row=zhu_ce_taskRow, column=1).value = dateData01
for index in range(0, len(zhu_ce_data)):
    zhu_ce_ws.cell(row=zhu_ce_taskRow, column=index + 3).value = zhu_ce_data[index]
    zhu_ce_ws.cell(row=zhu_ce_taskRow, column=index + 3).alignment = Alignment(horizontal='center', vertical='center')

# ----------zhu_ce end-----------

# ----------money_er start-----------
money_er_ws = wb.worksheets[3]
# 操作行数的位置
money_er_taskRow = len(money_er_ws.rows) + 1
money_er_taskColumn = len(money_er_ws.columns)

# print money_er_taskRow
# print money_er_taskColumn
# 转int
for index in range(0, len(money_er_data)):
    money_er_data[index] = int(money_er_data[index])

# 总和
for index in range(0, len(money_er_data)):
    money_er_data[len(money_er_data) - 2] += int(money_er_data[index])
# last value
money_er_data[len(money_er_data) - 1] = money_er_data[1] + money_er_data[2]

# 写数据
money_er_ws.cell(row=money_er_taskRow, column=1).value = dateData01
for index in range(0, len(money_er_data)):
    money_er_ws.cell(row=money_er_taskRow, column=index + 3).value = money_er_data[index]
    money_er_ws.cell(row=money_er_taskRow, column=index + 3).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

# ----------money_er end-----------

# ----------pay_for start-----------
pay_for_ws = wb.worksheets[4]
# 操作行数的位置
pay_for_taskRow = len(pay_for_ws.rows) + 1
pay_for_taskColumn = len(pay_for_ws.columns)

# print pay_for_taskRow
# print pay_for_taskColumn
# 转int
for index in range(0, len(pay_for_data)):
    pay_for_data[index] = float(pay_for_data[index])

# 总和
for index in range(0, len(pay_for_data)):
    pay_for_data[len(pay_for_data) - 2] += pay_for_data[index]
# last value
pay_for_data[len(pay_for_data) - 1] = pay_for_data[1] + pay_for_data[2]

# 写数据
pay_for_ws.cell(row=pay_for_taskRow, column=1).value = dateData01
for index in range(0, len(pay_for_data)):
    pay_for_ws.cell(row=pay_for_taskRow, column=index + 3).value = pay_for_data[index]
    pay_for_ws.cell(row=pay_for_taskRow, column=index + 3).alignment = Alignment(horizontal='center', vertical='center')

# ----------pay_for end-----------

# ------------总汇 start-------------
main_ws = wb.active
# 操作行数的位置
main_taskRow = len(main_ws.rows) + 1
main_taskColumn = 12
print main_taskRow
print main_taskColumn

select_file_open = open(select_file, 'r')
select_data = {}
index = 0
# 获取select数据
while True:
    line = select_file_open.readline()
    if line:
        select_data[index] = line.strip()
        index += 1
    else:
        break
print select_data

# print main_taskRow
# print main_taskColumn
main_data = ['', '', '', '', '', '', '', '', '', '', '', '']
main_data[0] = currentDay.strftime("%Y%m%d")
main_data[1] = select_data[0]
main_data[2] = select_data[1]
main_data[3] = select_data[2]
main_data[4] = pay_for_data[len(pay_for_data) - 2]
main_data[5] = money_er_data[len(money_er_data) - 2]
main_data[6] = round(float((main_data[4] / main_data[5])), 2)
main_data[7] = deng_lu_data[len(deng_lu_data) - 2]
main_data[8] = deng_lu_data[len(deng_lu_data) - 1]
main_data[9] = zhu_ce_data[len(zhu_ce_data) - 1]
main_data[10] = money_er_data[len(money_er_data) - 1]
main_data[11] = pay_for_data[len(pay_for_data) - 1]
print "汇总数据->\n%s" % main_data

for index in range(1, main_taskColumn + 1):
    main_ws.cell(row=main_taskRow, column=index).value = main_data[index - 1]
    main_ws.cell(row=main_taskRow, column=index).alignment = Alignment(horizontal='center', vertical='center')

# ------------总汇 end---------------

# 保存数据
wb.save(targetFilePath)

print "the task is successful~~~"
