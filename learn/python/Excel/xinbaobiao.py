# coding:utf-8
import os
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime
import sys

# 文件名 - 时间戳的准备
tempFileData = {}
currentDay = datetime.datetime.now()
yesterday = currentDay + datetime.timedelta(days=-1)
dateData = yesterday.strftime("%Y/%m/%d")

# /data/select/暴风影音Android_风秀_流量监测2016年8月15日.xls

# ------config-star-------
# 打开临时文件
tempFilePath = "/data/select/alic_xbb_" + yesterday.strftime("%Y-%m-%d") + ".xls"
sourceFilePath = "/data/select/暴风影音Android_风秀_流量监测" + yesterday.strftime("%Y年%m月%d日") + ".xlsx"
targetFilePath = "/data/select/暴风影音Android_风秀_流量监测" + currentDay.strftime("%Y年%m月%d日") + ".xlsx"
# ------config-end-------

# 判断文件是否存在
if not os.path.exists(tempFilePath):
    print tempFilePath + "文件不存在"
    print "Sorry! the task had stop~~~"
    sys.exit(0)

if not os.path.exists(sourceFilePath):
    print sourceFilePath + "文件不存在"
    print "Sorry! the task had stop~~~"
    sys.exit(0)  # 读取临时文件的数据

tempFile = open(tempFilePath, 'r')
# print tempFileData[1]
index = 1
while True:
    line = tempFile.readline()
    if line:
        tempFileData[index] = line
        # print tempFileData[index]
        # print line
        index += 1
    else:
        break
tempFile.close()
# print tempFileData

# 加载目标文件并追加
# print sourceFilePath

wb = load_workbook(sourceFilePath)

# 打开默认的工作表
default_ws = wb.active

# 获取工作的行数
taskRow = len(default_ws.rows) + 1
# 开始操作的列数
taskColumn = 12

# print taskRow

default_ws.cell(row=taskRow, column=1).value = dateData
default_ws.cell(row=taskRow, column=1).font = Font(size=10)
default_ws.cell(row=taskRow, column=1).alignment = Alignment(horizontal='center', vertical='center')
for index in range(0, len(tempFileData)):
    column = taskColumn + index
    # print column
    # print index + 1
    if index + 1 > len(tempFileData) - 3:
        tempFileData[index + 1] = round(float(tempFileData[index + 1]), 2)
    # print tempFileData[index + 1]
    # 数据与样式处理
    default_ws.cell(row=taskRow, column=column).value = tempFileData[index + 1]
    default_ws.cell(row=taskRow, column=column).font = Font(size=10)
    default_ws.cell(row=taskRow, column=column).alignment = Alignment(horizontal='center', vertical='center')

# # 保存数据
wb.save(targetFilePath)

print "the task is successful~~~"
